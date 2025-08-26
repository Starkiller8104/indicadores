# -*- coding: utf-8 -*-
# -*- coding: utf-8 -*-

# -*- coding: utf-8 -*-
"""
App Streamlit: Automatización de Indicadores (IMEMSA)
"""

import io, os, re, requests, pandas as pd, pytz, feedparser
from datetime import datetime
from urllib.parse import quote_plus
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import streamlit as st

# --------------------------
# Configuración de página (si ya tienes una, puedes conservarla)
# --------------------------
st.set_page_config(page_title="Automatización Indicadores", page_icon="📊", layout="centered")

# ---- Login (patch) ----
import os, pytz
def _get_app_password() -> str:
    try:
        return st.secrets["APP_PASSWORD"]
    except Exception:
        pass
    if os.getenv("APP_PASSWORD"):
        return os.getenv("APP_PASSWORD")
    return "imemsa79"

def _check_password() -> bool:
    if "auth_ok" not in st.session_state:
        st.session_state.auth_ok = False
    def _try_login():
        pw = st.session_state.get("password_input", "")
        st.session_state.auth_ok = (pw == _get_app_password())
        st.session_state.password_input = ""
    if st.session_state.auth_ok:
        return True
    st.title("🔒 Acceso restringido")
    st.text_input("Contraseña", type="password", key="password_input", on_change=_try_login, placeholder="Escribe tu contraseña…")
    st.stop()
# ---- /Login (patch) ----

_check_password()

TZ_MX = pytz.timezone("America/Mexico_City")

# --------------------------
# Utilidades existentes de tu app…
# (aquí va todo tu código original sin quitar nada;
#  solo se muestran funciones clave que ya traías)
# --------------------------
def safe_round(x, n):
    try:
        return round(float(x), n)
    except Exception:
        return None

def sie_opportuno(series_ids, banxico_token: str):
    if isinstance(series_ids, (list, tuple)):
        sid = ",".join(series_ids)
    else:
        sid = series_ids
    url = f"https://www.banxico.org.mx/SieAPIRest/service/v1/series/{sid}/datos/oportuno"
    headers = {"Bmx-Token": banxico_token.strip()}
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    data = r.json().get("bmx", {}).get("series", [])
    out = {}
    for s in data:
        try:
            out[s["idSerie"]] = float(str(s["datos"][0]["dato"]).replace(",", ""))
        except Exception:
            out[s["idSerie"]] = None
    return out

def fetch_tiie_from_dof():
    try:
        url = "https://sidof.segob.gob.mx/historicoIndicadores"
        r = requests.get(url, timeout=30); r.raise_for_status()
        text = " ".join(BeautifulSoup(r.text, "lxml").stripped_strings)
        def grab(pat):
            m = re.search(pat, text, flags=re.I)
            return float(m.group(1)) if m else None
        return {
            "tiie_28": safe_round(grab(r"TIIE\s*28.*?([0-9]+(?:\.[0-9]+)?)"),4),
            "tiie_91": safe_round(grab(r"TIIE\s*91.*?([0-9]+(?:\.[0-9]+)?)"),4),
            "tiie_182": safe_round(grab(r"TIIE\s*182.*?([0-9]+(?:\.[0-9]+)?)"),4),
        }
    except Exception:
        return {"tiie_28":None,"tiie_91":None,"tiie_182":None}

def cetes_sie(banxico_token: str):
    ids = ["SF43936", "SF43939", "SF43942", "SF43945"]
    mp  = {"SF43936":"28","SF43939":"91","SF43942":"182","SF43945":"364"}
    out = {k:None for k in mp.values()}
    data = sie_opportuno(ids, banxico_token)
    for k,v in data.items():
        out[mp[k]] = safe_round(v,4)
    return out

def fetch_uma_values():
    try:
        url = "https://www.inegi.org.mx/temas/uma/"
        r = requests.get(url, timeout=30); r.raise_for_status()
        txt = " ".join(BeautifulSoup(r.text, "lxml").stripped_strings)
        d = re.search(r"Diaria.*?([0-9]+(?:\.[0-9]+)?)", txt)
        m = re.search(r"Mensual.*?([0-9]+(?:\.[0-9]+)?)", txt)
        a = re.search(r"Anual.*?([0-9]+(?:\.[0-9]+)?)", txt)
        return (float(d.group(1)), float(m.group(1)), float(a.group(1)))
    except Exception:
        return (113.14, 3439.46, 41273.52)

# ---- build_news_bullets (patch si no existía) ----
def build_news_bullets(max_items=10):
    feeds = [
        "https://www.reuters.com/markets/americas/mexico/feed/?rpc=401&",
        "https://www.eleconomista.com.mx/rss/economia",
        "https://www.elfinanciero.com.mx/rss/finanzas/",
        "https://www.bloomberglinea.com/mexico/rss/",
    ]
    keywords = ["México","Banxico","inflación","tasa","TIIE","CETES","dólar","tipo de cambio","Pemex","FOMC","nearshoring"]
    rows = []
    for url in feeds:
        try:
            fp = feedparser.parse(url)
            for e in fp.entries[:40]:
                title = (e.get("title","") or "").strip()
                summary = (e.get("summary","") or "")
                link = (e.get("link","") or "").strip()
                txt = f"{title} {summary}".lower()
                if any(k.lower() in txt for k in keywords):
                    rows.append((e.get("published",""), title, link))
        except Exception:
            pass
    try:
        rows.sort(reverse=True, key=lambda x: x[0])
    except Exception:
        pass
    bullets = [f"• {t} — {l}" for _, t, l in rows[:max_items]]
    return "\n".join(bullets) if bullets else "Sin novedades (verifica conexión y RSS)."
# ---- /build_news_bullets ----

# --------------------------
# UI (uploader original + PATCH tokens)
# --------------------------
# (Mantengo tu UI original; solo agrego los controles de tokens y noticias.)
uploaded = st.file_uploader("Selecciona tu archivo .xlsx", type=["xlsx"])

# ---- Tokens editables (patch) ----
with st.sidebar.expander("🔑 Tokens de APIs"):
    st.caption("Se guardarán en la hoja **Token** del Excel resultante.")
    token_banxico_input = st.text_input("BANXICO_TOKEN", value="", type="password")
    token_inegi_input   = st.text_input("INEGI_TOKEN", value="", type="password")
# ---- /Tokens editables ----

# ---- Noticias: checkbox + vista previa (patch) ----
run_news = st.checkbox("📰 Incluir noticias financieras en la hoja \"Noticias\"", value=True)
try:
    if run_news:
        st.markdown("### 📰 Noticias (previa)")
        st.markdown(build_news_bullets(max_items=8).replace("•","-"))
except Exception:
    st.caption("No se pudieron cargar las noticias en la vista previa.")
# ---- /Noticias ----

# --------------------------
# Procesamiento (se conserva tu lógica, con PATCH de tokens y noticias)
# --------------------------
if uploaded:
    # (Tu flujo original de carga de Excel)
    raw = uploaded.getvalue()
    wb = load_workbook(io.BytesIO(raw), data_only=True)

    # Validación de hojas esperadas
    for hoja in ("Token","Indicadores","Noticias"):
        if hoja not in wb.sheetnames:
            st.error(f"Falta hoja {hoja}.")
            st.stop()

    ws_tok, ws_ind, ws_new = wb["Token"], wb["Indicadores"], wb["Noticias"]

    # ---- PATCH: tokens editables con persistencia ----
    BANXICO_TOKEN = (token_banxico_input.strip() if token_banxico_input.strip() else str(ws_tok["A2"].value or "").strip())
    INEGI_TOKEN   = (token_inegi_input.strip()   if token_inegi_input.strip()   else str(ws_tok["C2"].value or "").strip())
    if not BANXICO_TOKEN:
        st.error("Falta BANXICO_TOKEN (barra lateral o Token!A2).")
        st.stop()
    # Si capturaste nuevos, escríbelos al Excel generado:
    if token_banxico_input.strip():
        ws_tok["A2"] = token_banxico_input.strip()
    if token_inegi_input.strip():
        ws_tok["C2"] = token_inegi_input.strip()
    # ---- /PATCH tokens ----

    FECHA_HOY = datetime.now(TZ_MX).strftime("%d/%m/%Y")

    # (Tu lógica original de consultas y cálculos; ejemplos:)
    fx = sie_opportuno(["SF43718","SF46406","SF46410"], BANXICO_TOKEN)
    usd_mxn, jpy_mxn, eur_mxn = fx.get("SF43718"), fx.get("SF46406"), fx.get("SF46410")
    usd_jpy = (usd_mxn / jpy_mxn) if (usd_mxn and jpy_mxn) else None
    eur_usd = (eur_mxn / usd_mxn) if (eur_mxn and usd_mxn) else None

    tiie = fetch_tiie_from_dof()
    cetes = cetes_sie(BANXICO_TOKEN)
    udis = sie_opportuno(["SP68257"], BANXICO_TOKEN).get("SP68257")
    uma_diaria, uma_mensual, uma_anual = fetch_uma_values()

    # ---- PATCH: noticias al Excel ----
    if run_news:
        ws_new["A2"] = build_news_bullets(12)
    # ---- /PATCH noticias ----

    # (Tu mapeo original de celdas; ejemplo:)
    ws_ind["F7"], ws_ind["L7"], ws_ind["F32"], ws_ind["K32"] = FECHA_HOY, FECHA_HOY, FECHA_HOY, FECHA_HOY
    ws_ind["F10"] = safe_round(usd_mxn,4)
    ws_ind["F16"] = safe_round(jpy_mxn,6)
    ws_ind["F17"] = safe_round(usd_jpy,6)
    ws_ind["F21"] = safe_round(eur_mxn,6)
    ws_ind["F22"] = safe_round(eur_usd,6)

    ws_ind["L9"]  = safe_round(tiie.get("tiie_28"), 4)
    ws_ind["L10"] = safe_round(tiie.get("tiie_91"), 4)
    ws_ind["L11"] = safe_round(tiie.get("tiie_182"), 4)

    ws_ind["L15"] = safe_round(cetes.get("28"), 4)
    ws_ind["L16"] = safe_round(cetes.get("91"), 4)
    ws_ind["L17"] = safe_round(cetes.get("182"), 4)
    ws_ind["L18"] = safe_round(cetes.get("364"), 4)

    ws_ind["F33"] = safe_round(udis,6)
    ws_ind["K33"] = safe_round(uma_diaria,2)
    ws_ind["K34"] = safe_round(uma_mensual,2)
    ws_ind["K35"] = safe_round(uma_anual,2)

    # Exportar
    out = io.BytesIO(); wb.save(out); out.seek(0)
    st.download_button("⬇️ Descargar Excel actualizado", data=out,
                       file_name="Indicadores_actualizado.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
